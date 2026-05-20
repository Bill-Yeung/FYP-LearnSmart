#!/usr/bin/env python3
"""
Generate Data Dictionary Excel file from DDL SQL file for FYP Report.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re

def parse_ddl(ddl_content: str) -> list[dict]:
    """Parse DDL content and extract table definitions."""
    tables = []

    # Find all CREATE TABLE statements
    table_pattern = r'CREATE TABLE (?:IF NOT EXISTS )?(\w+)\s*\((.*?)\);'
    matches = re.findall(table_pattern, ddl_content, re.DOTALL | re.IGNORECASE)

    for table_name, columns_def in matches:
        columns = parse_columns(table_name, columns_def)
        tables.append({
            'name': table_name,
            'columns': columns
        })

    return tables

def parse_columns(table_name: str, columns_def: str) -> list[dict]:
    """Parse column definitions from a CREATE TABLE statement."""
    columns = []

    # Split by commas but handle nested parentheses
    lines = split_columns(columns_def)

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Skip constraints
        if line.upper().startswith(('PRIMARY KEY', 'FOREIGN KEY', 'UNIQUE', 'CHECK', 'CONSTRAINT')):
            continue

        # Parse column definition
        col_info = parse_column_line(line, table_name)
        if col_info:
            columns.append(col_info)

    return columns

def split_columns(columns_def: str) -> list[str]:
    """Split column definitions handling nested parentheses."""
    result = []
    current = []
    depth = 0

    for char in columns_def:
        if char == '(':
            depth += 1
            current.append(char)
        elif char == ')':
            depth -= 1
            current.append(char)
        elif char == ',' and depth == 0:
            result.append(''.join(current))
            current = []
        else:
            current.append(char)

    if current:
        result.append(''.join(current))

    return result

def parse_column_line(line: str, table_name: str) -> dict | None:
    """Parse a single column definition line."""
    # Match column name and type
    match = re.match(r'^(\w+)\s+(.+)$', line.strip(), re.IGNORECASE)
    if not match:
        return None

    col_name = match.group(1)
    rest = match.group(2)

    # Skip if it looks like a constraint
    if col_name.upper() in ('PRIMARY', 'FOREIGN', 'UNIQUE', 'CHECK', 'CONSTRAINT', 'INDEX'):
        return None

    # Parse data type
    data_type, data_length = parse_data_type(rest)

    # Check for constraints
    is_pk = 'PRIMARY KEY' in rest.upper()
    is_fk = 'REFERENCES' in rest.upper()
    is_nullable = 'NOT NULL' not in rest.upper() and not is_pk

    # Extract default value
    default_value = extract_default(rest)

    # Extract check constraint
    constraint = extract_constraint(rest)

    # Get min/max values based on type and constraint
    min_val, max_val = get_min_max(data_type, data_length, constraint)

    # Get description and example
    description = get_description(table_name, col_name, data_type)
    example = get_example(table_name, col_name, data_type, constraint)

    return {
        'table_name': table_name,
        'column_name': col_name,
        'data_type': data_type,
        'data_length': data_length,
        'description': description,
        'min_value': min_val,
        'max_value': max_val,
        'default_value': default_value,
        'pk_fk': 'PK' if is_pk else ('FK' if is_fk else '-'),
        'constraint': constraint,
        'nullable': 'Yes' if is_nullable else 'No',
        'example': example
    }

def parse_data_type(type_str: str) -> tuple[str, str]:
    """Extract data type and length from type string."""
    # Common patterns
    match = re.match(r'^(VARCHAR|CHAR|NUMERIC|DECIMAL|INT|INTEGER|BIGINT|SMALLINT|FLOAT|DOUBLE|TEXT|UUID|BOOLEAN|TIMESTAMP|TIMESTAMPTZ|DATE|TIME|SERIAL|BYTEA|INET|JSONB|JSON)\s*(?:\(([^)]+)\))?', type_str.upper())

    if match:
        data_type = match.group(1)
        length = match.group(2) if match.group(2) else '-'

        # Clean up type names
        type_mapping = {
            'VARCHAR': 'VARCHAR',
            'CHAR': 'CHAR',
            'TEXT': 'TEXT',
            'INTEGER': 'INTEGER',
            'INT': 'INTEGER',
            'BIGINT': 'BIGINT',
            'SMALLINT': 'SMALLINT',
            'SERIAL': 'SERIAL',
            'UUID': 'UUID',
            'BOOLEAN': 'BOOLEAN',
            'TIMESTAMP': 'TIMESTAMP',
            'TIMESTAMPTZ': 'TIMESTAMPTZ',
            'DATE': 'DATE',
            'TIME': 'TIME',
            'NUMERIC': 'NUMERIC',
            'DECIMAL': 'DECIMAL',
            'FLOAT': 'FLOAT',
            'DOUBLE': 'DOUBLE',
            'JSONB': 'JSONB',
            'JSON': 'JSON',
            'BYTEA': 'BYTEA',
            'INET': 'INET'
        }
        return type_mapping.get(data_type, data_type), length

    # Check for array types
    if '[]' in type_str:
        base_type = re.match(r'^(\w+)', type_str)
        if base_type:
            return f"{base_type.group(1).upper()}[]", '-'

    # Check for custom types (enums)
    first_word = type_str.split()[0].upper()
    return first_word, '-'

def extract_default(type_str: str) -> str:
    """Extract default value from column definition."""
    match = re.search(r"DEFAULT\s+(.+?)(?:\s+(?:NOT NULL|NULL|CHECK|REFERENCES|PRIMARY|UNIQUE|CONSTRAINT)|$)", type_str, re.IGNORECASE)
    if match:
        default = match.group(1).strip()
        # Clean up the default value
        default = re.sub(r'\s+', ' ', default)
        if len(default) > 50:
            default = default[:47] + '...'
        return default
    return '-'

def extract_constraint(type_str: str) -> str:
    """Extract CHECK constraint from column definition."""
    match = re.search(r"CHECK\s*\((.+?)\)(?:\s|,|$)", type_str, re.IGNORECASE)
    if match:
        constraint = match.group(1).strip()
        if len(constraint) > 80:
            constraint = constraint[:77] + '...'
        return constraint

    # Check for REFERENCES (foreign key)
    match = re.search(r"REFERENCES\s+(\w+)\((\w+)\)", type_str, re.IGNORECASE)
    if match:
        return f"FK -> {match.group(1)}({match.group(2)})"

    return '-'

def get_min_max(data_type: str, data_length: str, constraint: str) -> tuple[str, str]:
    """Determine min/max values based on data type and constraints."""
    # Check constraint for explicit bounds
    if constraint != '-':
        # Look for BETWEEN x AND y
        match = re.search(r'BETWEEN\s+(\d+)\s+AND\s+(\d+)', constraint, re.IGNORECASE)
        if match:
            return match.group(1), match.group(2)

        # Look for >= x or <= y
        match_min = re.search(r'>=\s*(\d+)', constraint)
        match_max = re.search(r'<=\s*(\d+)', constraint)
        min_val = match_min.group(1) if match_min else '-'
        max_val = match_max.group(1) if match_max else '-'
        if min_val != '-' or max_val != '-':
            return min_val, max_val

    # Based on data type - only numeric types have min/max
    if data_type == 'INTEGER':
        return '-2147483648', '2147483647'
    elif data_type == 'SMALLINT':
        return '-32768', '32767'
    elif data_type == 'BIGINT':
        return '-9223372036854775808', '9223372036854775807'
    elif data_type in ('NUMERIC', 'DECIMAL') and data_length != '-':
        parts = data_length.split(',')
        if len(parts) == 2:
            precision, scale = int(parts[0]), int(parts[1])
            max_int_digits = precision - scale
            max_val = '9' * max_int_digits + '.' + '9' * scale
            return f"-{max_val}", max_val
    elif data_type == 'BOOLEAN':
        return 'false', 'true'

    # VARCHAR, TEXT, CHAR, UUID, TIMESTAMP, etc. don't have min/max
    return '-', '-'

def get_description(table_name: str, col_name: str, data_type: str) -> str:
    """Generate description based on table and column context."""
    descriptions = {
        # Users table
        ('users', 'id'): 'Unique identifier for the user account',
        ('users', 'username'): 'Unique username for login and display',
        ('users', 'email'): 'User email address for authentication and notifications',
        ('users', 'password_hash'): 'Bcrypt hashed password for secure authentication',
        ('users', 'role'): 'User role determining access permissions (student/teacher/admin)',
        ('users', 'display_name'): 'Display name shown in the user interface',
        ('users', 'preferred_language'): 'User preferred language for interface localization',
        ('users', 'is_active'): 'Whether the user account is active and can log in',
        ('users', 'email_verified'): 'Whether the user has verified their email address',
        ('users', 'oauth_provider'): 'Third-party OAuth provider name (Google, GitHub, etc.)',
        ('users', 'oauth_id'): 'User ID from the OAuth provider',
        ('users', 'created_at'): 'Timestamp when the user account was created',
        ('users', 'updated_at'): 'Timestamp when the user account was last updated',
        ('users', 'last_login'): 'Timestamp of the user\'s most recent login',

        # User sessions
        ('user_sessions', 'id'): 'Unique identifier for the session',
        ('user_sessions', 'user_id'): 'Reference to the user who owns this session',
        ('user_sessions', 'token_hash'): 'Hashed JWT refresh token for session validation',
        ('user_sessions', 'ip_address'): 'IP address from which the session was created',
        ('user_sessions', 'user_agent'): 'Browser/client user agent string',
        ('user_sessions', 'created_at'): 'Timestamp when the session was created',
        ('user_sessions', 'expires_at'): 'Timestamp when the session expires',
        ('user_sessions', 'last_activity'): 'Timestamp of last activity in this session',

        # User profiles
        ('user_profiles', 'user_id'): 'Reference to the user this profile belongs to',
        ('user_profiles', 'bio'): 'User biography or self-description',
        ('user_profiles', 'avatar_url'): 'URL to the user\'s profile picture',
        ('user_profiles', 'organization'): 'Organization or institution the user belongs to',
        ('user_profiles', 'department'): 'Department within the organization',
        ('user_profiles', 'level'): 'Academic or professional level',
        ('user_profiles', 'personal_interests'): 'Array of user\'s personal interests for personalization',
        ('user_profiles', 'timezone'): 'User\'s timezone for scheduling features',
        ('user_profiles', 'notification_preferences'): 'JSON configuration for notification settings',
        ('user_profiles', 'privacy_settings'): 'JSON configuration for privacy preferences',
        ('user_profiles', 'domain_level'): 'User\'s current domain knowledge level',
        ('user_profiles', 'difficulty_preference'): 'Preferred difficulty level for learning content',
        ('user_profiles', 'ai_assistance_level'): 'Preferred level of AI assistance',
        ('user_profiles', 'total_play_time_minutes'): 'Total time spent in gamified learning activities',
        ('user_profiles', 'scripts_completed'): 'Number of script-kill game scripts completed',
        ('user_profiles', 'study_preferences'): 'JSON object with study preference settings',
        ('user_profiles', 'learning_style'): 'JSON object with learning style scores (visual, auditory, etc.)',
        ('user_profiles', 'updated_at'): 'Timestamp when the profile was last updated',

        # User activity log
        ('user_activity_log', 'id'): 'Unique identifier for the activity log entry',
        ('user_activity_log', 'user_id'): 'Reference to the user who performed the action',
        ('user_activity_log', 'action_type'): 'Type of action performed (login, upload, view, etc.)',
        ('user_activity_log', 'resource_type'): 'Type of resource the action was performed on',
        ('user_activity_log', 'resource_id'): 'ID of the resource the action was performed on',
        ('user_activity_log', 'details'): 'Additional JSON details about the action',
        ('user_activity_log', 'ip_address'): 'IP address from which the action was performed',
        ('user_activity_log', 'created_at'): 'Timestamp when the action occurred',

        # Concepts
        ('concepts', 'id'): 'Unique identifier for the concept',
        ('concepts', 'concept_type'): 'Type of concept (definition, procedure, example, etc.)',
        ('concepts', 'difficulty_level'): 'Difficulty level of the concept',
        ('concepts', 'estimated_study_time_minutes'): 'Estimated time to study this concept',
        ('concepts', 'formula_latex'): 'LaTeX formula if the concept includes mathematical formulas',
        ('concepts', 'base_form'): 'Base form of the concept for disambiguation',
        ('concepts', 'created_by'): 'Reference to the user who created this concept',
        ('concepts', 'is_system_generated'): 'Whether the concept was auto-generated by AI',
        ('concepts', 'is_public'): 'Whether the concept is publicly visible',
        ('concepts', 'qdrant_synced_at'): 'Timestamp when synced to Qdrant vector database',
        ('concepts', 'embedding_model'): 'Name of the embedding model used for vectorization',
        ('concepts', 'created_at'): 'Timestamp when the concept was created',
        ('concepts', 'updated_at'): 'Timestamp when the concept was last updated',
        ('concepts', 'version'): 'Version number for tracking concept revisions',

        # Concept translations
        ('concept_translations', 'id'): 'Unique identifier for the translation entry',
        ('concept_translations', 'concept_id'): 'Reference to the concept being translated',
        ('concept_translations', 'language'): 'Language code of the translation (en, zh, etc.)',
        ('concept_translations', 'title'): 'Translated title of the concept',
        ('concept_translations', 'description'): 'Translated description with inline citations',
        ('concept_translations', 'keywords'): 'Array of keywords for search and categorization',
        ('concept_translations', 'formula_plain_text'): 'Plain text version of formula for accessibility',
        ('concept_translations', 'is_primary'): 'Whether this is the primary/original language',
        ('concept_translations', 'created_by'): 'Reference to user who created this translation',
        ('concept_translations', 'translation_quality'): 'Quality indicator (source, llm, user_verified)',
        ('concept_translations', 'translation_date'): 'Timestamp when the translation was created',

        # Sources (documents)
        ('sources', 'id'): 'Unique identifier for the uploaded document',
        ('sources', 'document_name'): 'Original filename of the uploaded document',
        ('sources', 'document_path'): 'Storage path of the document file',
        ('sources', 'document_type'): 'Type of document (pdf, word, excel, etc.)',
        ('sources', 'language'): 'Primary language of the document content',
        ('sources', 'author'): 'Author of the document',
        ('sources', 'publication_year'): 'Year the document was published',
        ('sources', 'uploaded_by'): 'Reference to the user who uploaded the document',
        ('sources', 'is_public'): 'Whether the document is publicly accessible',
        ('sources', 'checksum'): 'SHA-256 checksum for duplicate detection',
        ('sources', 'processing_status'): 'Current processing status (pending, processing, completed, failed)',
        ('sources', 'processing_error'): 'Error message if processing failed',
        ('sources', 'processing_started_at'): 'Timestamp when document processing started',
        ('sources', 'processing_completed_at'): 'Timestamp when document processing completed',
        ('sources', 'concepts_extracted'): 'Number of concepts extracted from the document',
        ('sources', 'relationships_extracted'): 'Number of relationships extracted from the document',
        ('sources', 'ai_summary'): 'AI-generated summary of the document content',
        ('sources', 'ai_summary_generated_at'): 'Timestamp when AI summary was generated',
        ('sources', 'deleted_at'): 'Timestamp when soft-deleted (null if active)',
        ('sources', 'uploaded_at'): 'Timestamp when the document was uploaded',

        # Relationships
        ('relationships', 'id'): 'Unique identifier for the relationship',
        ('relationships', 'relationship_type'): 'Type of relationship between concepts',
        ('relationships', 'suggested_relationship_type'): 'AI-suggested relationship type if custom',
        ('relationships', 'direction'): 'Direction of relationship (unidirectional/bidirectional)',
        ('relationships', 'strength'): 'Strength of the relationship (0.0 to 1.0)',
        ('relationships', 'created_at'): 'Timestamp when the relationship was created',
        ('relationships', 'version'): 'Version number for tracking revisions',

        # Flashcards
        ('flashcards', 'id'): 'Unique identifier for the flashcard',
        ('flashcards', 'user_id'): 'Reference to the user who owns this flashcard',
        ('flashcards', 'concept_id'): 'Reference to the related concept (optional)',
        ('flashcards', 'taxonomy_node_id'): 'Reference to the taxonomy category',
        ('flashcards', 'front_content'): 'Content displayed on the front of the flashcard',
        ('flashcards', 'back_content'): 'Content displayed on the back (answer)',
        ('flashcards', 'card_type'): 'Type of card (standard or multiple choice)',
        ('flashcards', 'tips'): 'JSON array of hints and tips for learning',
        ('flashcards', 'content_metadata'): 'Additional metadata about the flashcard content',
        ('flashcards', 'source_type'): 'How the flashcard was created (manual, import, generated)',
        ('flashcards', 'is_archived'): 'Whether the flashcard is archived',
        ('flashcards', 'created_at'): 'Timestamp when the flashcard was created',

        # Flashcard schedules (spaced repetition)
        ('flashcard_schedules', 'flashcard_id'): 'Reference to the flashcard',
        ('flashcard_schedules', 'user_id'): 'Reference to the user',
        ('flashcard_schedules', 'algorithm'): 'Spaced repetition algorithm (simple, sm2, fsrs)',
        ('flashcard_schedules', 'state'): 'Current learning state (new, learning, review, relearning)',
        ('flashcard_schedules', 'due_date'): 'Next scheduled review date',
        ('flashcard_schedules', 'last_review_date'): 'Date of the last review',
        ('flashcard_schedules', 'interval_days'): 'Current review interval in days',
        ('flashcard_schedules', 'reps'): 'Total number of reviews completed',
        ('flashcard_schedules', 'ease_factor'): 'SM-2 ease factor for interval calculation',
        ('flashcard_schedules', 'stability'): 'FSRS stability parameter',
        ('flashcard_schedules', 'difficulty'): 'FSRS difficulty parameter',
        ('flashcard_schedules', 'topic_cached'): 'Cached topic name for quick display',

        # Tags
        ('tags', 'id'): 'Unique identifier for the tag',
        ('tags', 'user_id'): 'Reference to the user who created the tag',
        ('tags', 'name'): 'Display name of the tag',
        ('tags', 'url_id'): 'URL-safe identifier for the tag',
        ('tags', 'description'): 'Description of what the tag represents',
        ('tags', 'color'): 'Hex color code for visual display',
        ('tags', 'icon'): 'Icon identifier for the tag',
        ('tags', 'is_system'): 'Whether this is a system-defined tag',
        ('tags', 'usage_count'): 'Number of times the tag has been applied',
        ('tags', 'created_at'): 'Timestamp when the tag was created',
        ('tags', 'updated_at'): 'Timestamp when the tag was last updated',

        # Communities
        ('communities', 'id'): 'Unique identifier for the community',
        ('communities', 'name'): 'Display name of the community',
        ('communities', 'url_id'): 'URL-safe identifier for the community',
        ('communities', 'description'): 'Description of the community purpose',
        ('communities', 'community_type'): 'Type of community (public, private, invite_only, course_based)',
        ('communities', 'max_members'): 'Maximum number of members allowed',
        ('communities', 'avatar_url'): 'URL to the community avatar image',
        ('communities', 'banner_url'): 'URL to the community banner image',
        ('communities', 'color_theme'): 'Theme color for the community',
        ('communities', 'features_enabled'): 'JSON config of enabled community features',
        ('communities', 'member_count'): 'Current number of community members',
        ('communities', 'resource_count'): 'Number of shared resources in the community',
        ('communities', 'activity_score'): 'Calculated activity score for ranking',
        ('communities', 'created_by'): 'Reference to the user who created the community',
        ('communities', 'created_at'): 'Timestamp when the community was created',
        ('communities', 'updated_at'): 'Timestamp when the community was last updated',

        # Game sessions (Script-Kill mode)
        ('game_sessions', 'session_id'): 'Unique identifier for the game session',
        ('game_sessions', 'id'): 'Reference to the player user',
        ('game_sessions', 'script_id'): 'Reference to the game script being played',
        ('game_sessions', 'session_type'): 'Type of session (solo, demo, test)',
        ('game_sessions', 'current_scene'): 'Current scene identifier in the game',
        ('game_sessions', 'game_progress'): 'JSON object storing game progress data',
        ('game_sessions', 'collected_evidence'): 'JSON array of collected evidence items',
        ('game_sessions', 'decisions_made'): 'JSON array of decisions made during gameplay',
        ('game_sessions', 'progress_percentage'): 'Overall game completion percentage',
        ('game_sessions', 'time_spent_minutes'): 'Total time spent in the session',
        ('game_sessions', 'achieved_ending'): 'Identifier of the ending achieved',
        ('game_sessions', 'ending_score'): 'Score for the achieved ending',
        ('game_sessions', 'status'): 'Current session status (active, paused, completed, abandoned)',
        ('game_sessions', 'started_at'): 'Timestamp when the session started',
        ('game_sessions', 'last_updated_at'): 'Timestamp of last session update',
        ('game_sessions', 'completed_at'): 'Timestamp when the session was completed',

        # Generated scripts (Script-Kill mode)
        ('generated_scripts', 'script_id'): 'Unique identifier for the generated script',
        ('generated_scripts', 'template_id'): 'Reference to the template used for generation',
        ('generated_scripts', 'generation_parameters'): 'JSON parameters used during generation',
        ('generated_scripts', 'script_title'): 'Title of the generated script',
        ('generated_scripts', 'script_content'): 'JSON content of the script (scenes, dialogues, etc.)',
        ('generated_scripts', 'script_summary'): 'Brief summary of the script storyline',
        ('generated_scripts', 'generation_method'): 'How the script was generated (manual, ai_assisted, ai_generated)',
        ('generated_scripts', 'ai_model_used'): 'AI model used for generation',
        ('generated_scripts', 'generation_prompt'): 'Prompt used for AI generation',
        ('generated_scripts', 'learning_points'): 'JSON array of learning objectives covered',
        ('generated_scripts', 'estimated_duration'): 'Estimated play duration in minutes',
        ('generated_scripts', 'validation_status'): 'Validation status (pending, validating, passed, failed)',
        ('generated_scripts', 'validation_score'): 'Score from validation process',
        ('generated_scripts', 'validation_notes'): 'Notes from the validation process',
        ('generated_scripts', 'is_active'): 'Whether the script is active and playable',
        ('generated_scripts', 'play_count'): 'Number of times the script has been played',
        ('generated_scripts', 'generated_at'): 'Timestamp when the script was generated',
        ('generated_scripts', 'last_played_at'): 'Timestamp when the script was last played',

        # Learning analytics
        ('learning_analytics', 'analytics_id'): 'Unique identifier for the analytics record',
        ('learning_analytics', 'id'): 'Reference to the user',
        ('learning_analytics', 'session_id'): 'Reference to the game session',
        ('learning_analytics', 'script_id'): 'Reference to the script played',
        ('learning_analytics', 'knowledge_points_covered'): 'JSON array of knowledge points encountered',
        ('learning_analytics', 'knowledge_mastery_score'): 'Score for knowledge mastery demonstrated',
        ('learning_analytics', 'reasoning_accuracy'): 'Accuracy of reasoning during gameplay',
        ('learning_analytics', 'puzzle_success_rate'): 'Success rate on puzzle challenges',
        ('learning_analytics', 'evidence_collection_rate'): 'Rate of evidence collection',
        ('learning_analytics', 'decision_quality_score'): 'Quality score for decisions made',
        ('learning_analytics', 'hints_requested'): 'Number of hints requested',
        ('learning_analytics', 'time_spent_on_knowledge'): 'Time spent on knowledge-related activities',
        ('learning_analytics', 'overall_score'): 'Overall performance score',
        ('learning_analytics', 'learning_efficiency'): 'Calculated learning efficiency metric',
        ('learning_analytics', 'improvement_suggestions'): 'JSON array of improvement suggestions',
        ('learning_analytics', 'recommended_next_scripts'): 'JSON array of recommended scripts',
        ('learning_analytics', 'analyzed_at'): 'Timestamp when analytics were calculated',
    }

    # Try to find specific description
    key = (table_name.lower(), col_name.lower())
    if key in descriptions:
        return descriptions[key]

    # Generate generic description based on column name patterns
    col_lower = col_name.lower()

    if col_lower == 'id':
        return f'Unique identifier for {table_name.replace("_", " ")} record'
    elif col_lower.endswith('_id'):
        ref_table = col_lower[:-3].replace('_', ' ')
        return f'Foreign key reference to {ref_table}'
    elif col_lower.endswith('_at'):
        action = col_lower[:-3].replace('_', ' ')
        return f'Timestamp when {action} occurred'
    elif col_lower.startswith('is_'):
        flag = col_lower[3:].replace('_', ' ')
        return f'Boolean flag indicating if {flag}'
    elif col_lower.endswith('_count'):
        item = col_lower[:-6].replace('_', ' ')
        return f'Count of {item}'
    elif col_lower.endswith('_score'):
        item = col_lower[:-6].replace('_', ' ')
        return f'Score value for {item}'
    elif col_lower.endswith('_url'):
        item = col_lower[:-4].replace('_', ' ')
        return f'URL for {item}'
    elif col_lower == 'name':
        return f'Name of the {table_name.replace("_", " ")}'
    elif col_lower == 'description':
        return f'Description of the {table_name.replace("_", " ")}'
    elif col_lower == 'status':
        return f'Current status of the {table_name.replace("_", " ")}'
    elif col_lower == 'title':
        return f'Title of the {table_name.replace("_", " ")}'
    elif col_lower == 'content':
        return f'Content of the {table_name.replace("_", " ")}'
    elif col_lower == 'metadata':
        return 'JSON object containing additional metadata'
    elif data_type == 'JSONB':
        return f'JSON data for {col_name.replace("_", " ")}'

    return f'{col_name.replace("_", " ").title()} field'

def get_example(table_name: str, col_name: str, data_type: str, constraint: str) -> str:
    """Generate example value based on table, column, and data type."""
    examples = {
        # Common patterns
        ('users', 'id'): '550e8400-e29b-41d4-a716-446655440000',
        ('users', 'username'): 'john_doe',
        ('users', 'email'): 'john.doe@example.com',
        ('users', 'password_hash'): '$2b$12$LQv3c1yq...',
        ('users', 'role'): 'student',
        ('users', 'display_name'): 'John Doe',
        ('users', 'preferred_language'): 'en',
        ('users', 'is_active'): 'true',
        ('users', 'email_verified'): 'true',
        ('users', 'oauth_provider'): 'google',
        ('users', 'oauth_id'): '118234567890123456789',

        ('user_profiles', 'bio'): 'Computer Science student interested in AI',
        ('user_profiles', 'avatar_url'): 'https://cdn.example.com/avatars/user123.jpg',
        ('user_profiles', 'organization'): 'Hong Kong Institute of Vocational Education',
        ('user_profiles', 'department'): 'Information Technology',
        ('user_profiles', 'level'): 'Year 2',
        ('user_profiles', 'timezone'): 'Asia/Hong_Kong',
        ('user_profiles', 'domain_level'): 'intermediate',
        ('user_profiles', 'difficulty_preference'): 'adaptive',
        ('user_profiles', 'ai_assistance_level'): 'moderate',
        ('user_profiles', 'total_play_time_minutes'): '120',
        ('user_profiles', 'scripts_completed'): '5',

        ('concepts', 'concept_type'): 'definition',
        ('concepts', 'difficulty_level'): 'intermediate',
        ('concepts', 'estimated_study_time_minutes'): '15',
        ('concepts', 'formula_latex'): 'E = mc^2',
        ('concepts', 'base_form'): 'algorithm',
        ('concepts', 'embedding_model'): 'BAAI/bge-small-zh-v1.5',
        ('concepts', 'version'): '1',

        ('concept_translations', 'language'): 'en',
        ('concept_translations', 'title'): 'Binary Search Algorithm',
        ('concept_translations', 'translation_quality'): 'user_verified',

        ('sources', 'document_name'): 'Data Structures Lecture Notes.pdf',
        ('sources', 'document_path'): '/uploads/2024/01/abc123.pdf',
        ('sources', 'document_type'): 'pdf',
        ('sources', 'language'): 'en',
        ('sources', 'author'): 'Dr. Smith',
        ('sources', 'publication_year'): '2024',
        ('sources', 'processing_status'): 'completed',
        ('sources', 'concepts_extracted'): '25',
        ('sources', 'relationships_extracted'): '18',

        ('relationships', 'relationship_type'): 'prerequisite_of',
        ('relationships', 'direction'): 'unidirectional',
        ('relationships', 'strength'): '0.85',

        ('flashcards', 'front_content'): 'What is the time complexity of binary search?',
        ('flashcards', 'back_content'): 'O(log n)',
        ('flashcards', 'card_type'): 'standard',
        ('flashcards', 'source_type'): 'note_generated',

        ('flashcard_schedules', 'algorithm'): 'sm2',
        ('flashcard_schedules', 'state'): 'review',
        ('flashcard_schedules', 'interval_days'): '7.5',
        ('flashcard_schedules', 'reps'): '5',
        ('flashcard_schedules', 'ease_factor'): '2.5',
        ('flashcard_schedules', 'stability'): '10.5',
        ('flashcard_schedules', 'difficulty'): '0.3',

        ('tags', 'name'): 'Programming',
        ('tags', 'url_id'): 'programming',
        ('tags', 'color'): '#3B82F6',
        ('tags', 'icon'): 'code',
        ('tags', 'usage_count'): '42',

        ('communities', 'name'): 'CS Study Group',
        ('communities', 'url_id'): 'cs-study-group',
        ('communities', 'community_type'): 'public',
        ('communities', 'max_members'): '100',
        ('communities', 'color_theme'): '#4F46E5',
        ('communities', 'member_count'): '45',
        ('communities', 'activity_score'): '850',

        ('game_sessions', 'session_type'): 'solo',
        ('game_sessions', 'current_scene'): 'library_investigation',
        ('game_sessions', 'progress_percentage'): '65',
        ('game_sessions', 'time_spent_minutes'): '45',
        ('game_sessions', 'achieved_ending'): 'true_ending',
        ('game_sessions', 'ending_score'): '92.5',
        ('game_sessions', 'status'): 'completed',

        ('generated_scripts', 'script_title'): 'The Missing Algorithm',
        ('generated_scripts', 'generation_method'): 'ai_assisted',
        ('generated_scripts', 'ai_model_used'): 'gpt-4',
        ('generated_scripts', 'estimated_duration'): '60',
        ('generated_scripts', 'validation_status'): 'passed',
        ('generated_scripts', 'validation_score'): '95.00',
        ('generated_scripts', 'play_count'): '128',

        ('learning_analytics', 'knowledge_mastery_score'): '85.50',
        ('learning_analytics', 'reasoning_accuracy'): '78.25',
        ('learning_analytics', 'puzzle_success_rate'): '90.00',
        ('learning_analytics', 'evidence_collection_rate'): '92.50',
        ('learning_analytics', 'decision_quality_score'): '88.75',
        ('learning_analytics', 'hints_requested'): '3',
        ('learning_analytics', 'overall_score'): '87.25',
        ('learning_analytics', 'learning_efficiency'): '1.25',
    }

    key = (table_name.lower(), col_name.lower())
    if key in examples:
        return examples[key]

    # Generate based on data type
    col_lower = col_name.lower()

    if data_type == 'UUID':
        return '550e8400-e29b-41d4-a716-446655440000'
    elif data_type == 'SERIAL':
        return '1'
    elif data_type == 'INTEGER':
        if 'year' in col_lower:
            return '2024'
        elif 'count' in col_lower or 'minutes' in col_lower:
            return '10'
        elif 'order' in col_lower or 'level' in col_lower:
            return '1'
        return '100'
    elif data_type == 'SMALLINT':
        return '1'
    elif data_type == 'BIGINT':
        return '1000000'
    elif data_type == 'BOOLEAN':
        return 'true'
    elif data_type in ('TIMESTAMP', 'TIMESTAMPTZ'):
        return '2024-01-15 10:30:00'
    elif data_type == 'DATE':
        return '2024-01-15'
    elif data_type == 'TIME':
        return '10:30:00'
    elif data_type in ('NUMERIC', 'DECIMAL', 'FLOAT', 'DOUBLE'):
        if 'score' in col_lower or 'rate' in col_lower:
            return '85.50'
        elif 'factor' in col_lower:
            return '2.5'
        elif 'percentage' in col_lower:
            return '75.5'
        return '100.00'
    elif data_type == 'VARCHAR':
        if 'email' in col_lower:
            return 'user@example.com'
        elif 'url' in col_lower:
            return 'https://example.com/path'
        elif 'name' in col_lower:
            return 'Sample Name'
        elif 'code' in col_lower:
            return 'ABC123'
        elif 'color' in col_lower:
            return '#3B82F6'
        elif 'language' in col_lower:
            return 'en'
        elif 'status' in col_lower:
            return 'active'
        elif 'type' in col_lower:
            return 'default'
        return 'Sample text value'
    elif data_type == 'TEXT':
        if 'description' in col_lower:
            return 'A detailed description of the item'
        elif 'content' in col_lower:
            return 'The main content text goes here'
        elif 'notes' in col_lower:
            return 'Additional notes and comments'
        return 'Long text content...'
    elif data_type == 'JSONB':
        if 'preferences' in col_lower:
            return '{"email": true, "push": false}'
        elif 'settings' in col_lower:
            return '{"theme": "dark", "language": "en"}'
        elif 'metadata' in col_lower:
            return '{"key": "value"}'
        return '{}'
    elif data_type == 'INET':
        return '192.168.1.1'
    elif data_type == 'BYTEA':
        return '<binary data>'
    elif '[]' in data_type:
        return "['item1', 'item2']"

    return 'Example value'

def create_excel(tables: list[dict], output_path: str):
    """Create Excel file with data dictionary."""
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"

    # Define headers (removed 'Table Name' column)
    headers = [
        'Data Item Name', 'Data Type', 'Data Length',
        'Data Description', 'Min Value', 'Max Value', 'Default Value',
        'PK or FK', 'Constraint', 'Nullable', 'Example'
    ]
    num_cols = len(headers)

    # Define column widths
    col_widths = [30, 15, 12, 60, 25, 25, 30, 10, 40, 10, 35]

    # Style definitions
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    table_name_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    table_name_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Write data
    row = 2
    alt_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    total_columns = 0

    for table in tables:
        # Add table name header row
        table_name = table['name']

        # Merge cells for table name row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value=f"Table: {table_name}")
        cell.fill = table_name_fill
        cell.font = table_name_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

        # Apply border to all merged cells
        for col in range(2, num_cols + 1):
            ws.cell(row=row, column=col).border = border

        ws.row_dimensions[row].height = 28
        row += 1

        # Write column data
        for idx, col_info in enumerate(table['columns']):
            # Alternate row colors
            fill = alt_fill if (idx % 2 == 0) else None

            values = [
                col_info['column_name'],
                col_info['data_type'],
                col_info['data_length'],
                col_info['description'],
                col_info['min_value'],
                col_info['max_value'],
                col_info['default_value'],
                col_info['pk_fk'],
                col_info['constraint'],
                col_info['nullable'],
                col_info['example']
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=str(value))
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if fill:
                    cell.fill = fill

            ws.row_dimensions[row].height = 25
            row += 1
            total_columns += 1

    # Set header row height
    ws.row_dimensions[1].height = 30

    # Save workbook
    wb.save(output_path)
    print(f"Data dictionary saved to: {output_path}")
    print(f"Total tables: {len(tables)}")
    print(f"Total columns: {total_columns}")

def main():
    # Read DDL file
    ddl_path = '/home/noelchan/CodeSpace/IVE/SE_Group10_FYP/backend/database/init_pg_ddl.sql'
    output_path = '/home/noelchan/CodeSpace/IVE/SE_Group10_FYP/Data_Dictionary.xlsx'

    with open(ddl_path, 'r', encoding='utf-8') as f:
        ddl_content = f.read()

    # Parse DDL
    tables = parse_ddl(ddl_content)

    # Create Excel
    create_excel(tables, output_path)

if __name__ == '__main__':
    main()
